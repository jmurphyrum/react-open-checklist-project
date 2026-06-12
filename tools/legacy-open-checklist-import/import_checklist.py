import argparse
import html
import re
import uuid
from html.parser import HTMLParser
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DATA_DIR = Path("data")


PROFILES = {
    "2025-26-topps-chrome-platinum": {
        "xlsx": "import/2025-Topps-Chrome-Platinum-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2025-26_Topps_Chrome_Platinum",
        "set_name": "2025-26 Topps Chrome Platinum Anniversary",
        "series": "Topps Chrome Platinum Anniversary",
        "season": "2025-26",
        "years": [2025, 2026],
        "release_date": "2026-06-05",
        "manufacturer": "Topps",
        "sheets": ["Base", "Variations", "Autographs", "Inserts"],
        "base_sections": {"Base"},
        "image_root": "https://example.com/cards/2025-26-topps-chrome-platinum",
        "set_image_url": "https://example.com/images/2025-26-topps-chrome-platinum.jpg",
    },
    "2026-bowman": {
        "xlsx": "import/2026-Bowman-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2026_Bowman#Checklist",
        "set_name": "2026 Bowman",
        "series": "2026 Bowman",
        "season": "2026",
        "years": [2026],
        "release_date": "2026-05-13",
        "manufacturer": "Topps",
        "sheets": ["Base", "Prospects", "Autographs", "Inserts"],
        "base_sections": {"Base Set"},
        "image_root": "https://example.com/cards/2026-bowman",
        "set_image_url": "https://example.com/images/2026-bowman.jpg",
        "section_print_runs": {
            "Chrome Prospects Packfractor Variation": 89,
            "Chrome Prospect Gold Ink Autographs": 15,
            "Chrome Prospect Packfractor Autographs": 89,
            "Draft Pick Pairings": 25,
            "Bowman Sterling Autographs": 150,
            "Electric Sluggers Autographs": 99,
            "Under The Radar Autographs": 99,
            "Power Chords Autographs": 99,
            "Ultimate Autograph Booklet": 10,
        },
        "section_estimated_print_runs": {
            "Bowman Sterling": 107800,
            "Electric Sluggers": 69625,
            "Under The Radar": 80900,
            "Power Chords": 29075,
        },
    },
    "2026-topps-series-2": {
        "xlsx": "import/2026-Topps-Series-2-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2026_Topps#Checklist",
        "set_name": "2026 Topps Series 2",
        "series": "2026 Topps Baseball",
        "series_number": 2,
        "season": "2026",
        "years": [2026],
        "release_date": "2026-06-11",
        "manufacturer": "Topps",
        "sheets": ["Base", "Variations", "Autographs", "Memorabilia", "Inserts"],
        "base_sections": {"Base"},
        "image_root": "https://example.com/cards/2026-topps-series-2",
        "set_image_url": "https://example.com/images/2026-topps-series-2.jpg",
        "section_print_runs": {
            "In The Name Relics": 1,
        },
        "section_relics": {
            "Major League Material",
            "City Connect Swatch Collection",
            "Real One Relics",
            "Rounding The Bases Relics",
            "1991 Topps Baseball Relics",
            "1991 Topps Baseball All-Star Relics",
            "In The Name Relics",
            "Topps Flagship Autograph Patches",
            "Heavy Lumber Autograph Relics",
            "Major League Materials Autographs",
            "Major Legaue Materials Dual Autographs",
            "City Connect Swatches Autographs",
            "Rounding The Bases Autographs",
        },
    },
    "2026-donruss": {
        "xlsx": "import/2026-Donruss-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2026_Donruss#Checklist",
        "set_name": "2026 Donruss",
        "series": "2026 Donruss",
        "season": "2026",
        "years": [2026],
        "release_date": "2026-05-27",
        "manufacturer": "Panini",
        "category": ["Baseball"],
        "sheets": ["Base", "Optic", "Autographs", "Memorabilia", "Inserts"],
        "base_sections": {"Base Set"},
        "image_root": "https://example.com/cards/2026-donruss",
        "set_image_url": "https://example.com/images/2026-donruss.jpg",
        "section_relics": {"Jersey Kings", "Prospect Jersey Kings", "Retro 1985 Materials"},
    },
    "2026-topps-chrome-black": {
        "xlsx": "import/2026-Topps-Chrome-Black-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2026_Topps_Chrome_Black#Checklist",
        "set_name": "2026 Topps Chrome Black",
        "series": "2026 Topps Chrome Black",
        "season": "2026",
        "years": [2026],
        "release_date": "2026-04-29",
        "manufacturer": "Topps",
        "sheets": ["Base", "Autographs", "Inserts"],
        "base_sections": {"Base Set"},
        "image_root": "https://example.com/cards/2026-topps-chrome-black",
        "set_image_url": "https://example.com/images/2026-topps-chrome-black.jpg",
    },
    "2025-26-topps-definitive-collection": {
        "xlsx": "import/2025-Topps-Definitive-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2025-26_Topps_Definitive_Collection#Checklist",
        "set_name": "2025-26 Topps Definitive Collection",
        "series": "Topps Definitive Collection",
        "season": "2025-26",
        "years": [2025, 2026],
        "release_date": "2026-04-29",
        "manufacturer": "Topps",
        "sheets": ["Base", "Autographs", "Autographed Relics", "Memorabilia Cards"],
        "base_sections": set(),
        "image_root": "https://example.com/cards/2025-26-topps-definitive-collection",
        "set_image_url": "https://example.com/images/2025-26-topps-definitive-collection.jpg",
        "all_cards_serial_numbered": True,
        "default_print_run": 50,
    },
    "2026-panini-prizm-stars-stripes": {
        "xlsx": "import/2026-Panini-Prizm-Stars-Stripes-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2026_Panini_Stars_%26_Stripes_Prizm#Checklist",
        "set_name": "2026 Panini Stars & Stripes Prizm",
        "series": "2026 Panini Stars & Stripes Prizm",
        "season": "2026",
        "years": [2026],
        "release_date": "2026-04-03",
        "manufacturer": "Panini",
        "category": ["USA Baseball"],
        "sheets": ["Base", "Autographs", "Memorabilia", "Inserts"],
        "base_sections": {"Base Set"},
        "default_team": "USA Baseball",
        "image_root": "https://example.com/cards/2026-panini-prizm-stars-stripes",
        "set_image_url": "https://example.com/images/2026-panini-prizm-stars-stripes.jpg",
        "section_relics": {"Jumbo Materials Flag", "Jumbo Materials Jersey", "Jumbo Materials Cap", "USA Materials"},
    },
    "2025-26-topps-transcendent": {
        "xlsx": "import/2025-Topps-Transcendent-Baseball-Checklist.xlsx",
        "url": "https://baseballcardpedia.com/index.php/2025-26_Topps_Transcendent#Checklist",
        "set_name": "2025-26 Topps Transcendent",
        "series": "Topps Transcendent",
        "season": "2025-26",
        "years": [2025, 2026],
        "release_date": "2026-04-01",
        "manufacturer": "Topps",
        "sheets": ["Base", "Autographs", "Memorabilia Cards", "Inserts"],
        "base_sections": {"Base Set"},
        "image_root": "https://example.com/cards/2025-26-topps-transcendent",
        "set_image_url": "https://example.com/images/2025-26-topps-transcendent.jpg",
        "section_relics": {
            "Transcendent Collection Patches",
            "Legendary Relics",
            "Transcendent Collection Dual Patches",
            "Transcendent Collection Triple Patches",
            "Transcendent Collection MLB Logo Patches",
            "Autographed Relics",
            "Autographed Patches",
            "Rookie Showcase Autographed Patches",
            "Jumbo Autographed Patches",
            "World Series Autographed Relics",
            "Autographed MLB Logo Patch",
            "Cut Signature Relic Book",
        },
    },
}


class DescriptionParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_heading = False
        self.in_description = False
        self.current_heading = None
        self.heading_parts = []
        self.description_parts = []
        self.descriptions = []

    def handle_starttag(self, tag, attrs):
        if tag in {"h1", "h2", "h3"}:
            self.in_heading = True
            self.heading_parts = []
        elif tag == "p" and self.current_heading == "Description":
            self.in_description = True
            self.description_parts = []

    def handle_endtag(self, tag):
        if self.in_heading and tag in {"h1", "h2", "h3"}:
            self.current_heading = clean_text("".join(self.heading_parts))
            self.in_heading = False
        elif tag == "p" and self.in_description:
            text = clean_text("".join(self.description_parts))
            if text:
                self.descriptions.append(text)
            self.in_description = False

    def handle_data(self, data):
        if self.in_heading:
            self.heading_parts.append(data)
        if self.in_description:
            self.description_parts.append(data)

    def handle_entityref(self, name):
        self.handle_data(html.unescape(f"&{name};"))

    def handle_charref(self, name):
        self.handle_data(html.unescape(f"&#{name};"))


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def clean_name(value):
    return clean_text(value).rstrip(",").strip()


def slugify(value):
    value = str(value).lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def yaml_quote(value):
    escaped = str(value).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def read_existing_uuid(path):
    if not path.exists():
        return None
    match = re.search(r'^uuid:\s*"?([0-9a-fA-F-]{36})"?\s*$', path.read_text(encoding="utf-8"), re.MULTILINE)
    return match.group(1) if match else None


def get_uuid(path):
    return read_existing_uuid(path) or str(uuid.uuid4())


def is_count_line(value):
    return bool(re.match(r"^\d+\s+(cards?|players|figures)$", clean_text(value), re.IGNORECASE))


def is_single_value_row(values):
    first = clean_text(values[0] if values else None)
    return all(clean_text(value) == "" for value in values[1:4])


def next_nonblank_first(rows, start_index):
    for row in rows[start_index:]:
        values = list(row) + [None] * 5
        first = clean_text(values[0])
        if first:
            return first
    return ""


def is_section_header(rows, index):
    values = list(rows[index]) + [None] * 5
    first = clean_text(values[0])
    if not first or is_count_line(first) or not is_single_value_row(values):
        return False
    return is_count_line(next_nonblank_first(rows, index + 1))


def parse_print_run(value):
    text = clean_text(value)
    match = re.match(r"^/(\d+)$", text)
    return int(match.group(1)) if match else None


def is_card_row(values, profile):
    name = clean_text(values[1] if len(values) > 1 else None)
    team = clean_text(values[2] if len(values) > 2 else None)
    if not name:
        return False
    if team and parse_print_run(team) is None:
        return True
    return bool(profile.get("default_team"))


def split_values(value):
    return [part.strip() for part in re.split(r"\s*/\s*", clean_name(value)) if part.strip()]


def team_for_subject(index, teams):
    if not teams:
        return ""
    if len(teams) == 1:
        return teams[0]
    if index < len(teams):
        return teams[index]
    return teams[-1]


def is_rookie(value):
    return clean_text(value).upper() == "RC"


def section_abbreviation(section):
    words = re.findall(r"[A-Za-z0-9]+", section)
    abbreviation = "".join(word[0] for word in words if word).upper()
    return abbreviation or "CARD"


def generated_number(section, count):
    return f"{section_abbreviation(section)}-{count}"


def read_cards_from_workbook(path, profile):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    missing_number_counts = {}

    for sheet_name in profile["sheets"]:
        ws = wb[sheet_name]
        worksheet_rows = list(ws.iter_rows(values_only=True))
        current_section = None
        for index, row in enumerate(worksheet_rows):
            values = list(row) + [None] * 5
            if is_section_header(worksheet_rows, index):
                current_section = clean_text(values[0])
                continue
            if not current_section or not is_card_row(values, profile):
                continue

            number = clean_text(values[0])
            generated = False
            if not number:
                missing_number_counts[current_section] = missing_number_counts.get(current_section, 0) + 1
                number = generated_number(current_section, missing_number_counts[current_section])
                generated = True

            names = split_values(values[1])
            print_run = next((parse_print_run(value) for value in values[2:6] if parse_print_run(value) is not None), None)
            raw_team = clean_text(values[2])
            teams = split_values(raw_team) if raw_team and parse_print_run(raw_team) is None else []
            if not teams and profile.get("default_team"):
                teams = [profile["default_team"]]
            rows.append(
                {
                    "number": number,
                    "generated_number": generated,
                    "subjects": [
                        {
                            "name": name,
                            "team": team_for_subject(index, teams),
                        }
                        for index, name in enumerate(names)
                    ],
                    "rookie_card": is_rookie(values[3]),
                    "serial_numbered": print_run is not None,
                    "print_run": print_run,
                    "section": current_section,
                    "source_sheet": sheet_name,
                    "note": clean_text(values[3]),
                }
            )

    return merge_duplicate_cards(rows)


def merge_duplicate_cards(rows):
    cards = []
    by_key = {}
    for row in rows:
        key = (row["section"], row["number"])
        if key not in by_key:
            row["subjects"] = list(row["subjects"])
            by_key[key] = row
            cards.append(row)
            continue
        existing = by_key[key]
        existing["subjects"].extend(row["subjects"])
        existing["rookie_card"] = existing["rookie_card"] or row["rookie_card"]
        existing["serial_numbered"] = existing.get("serial_numbered") or row.get("serial_numbered")
        existing["print_run"] = existing.get("print_run") or row.get("print_run")
    return cards


def fetch_description(url, fallback):
    request = Request(url, headers={"User-Agent": "open-checklist-project-importer/0.3"})
    with urlopen(request, timeout=30) as response:
        parser = DescriptionParser()
        parser.feed(response.read().decode("utf-8", errors="replace"))
    return parser.descriptions[0] if parser.descriptions else fallback


def card_filename(number, section, profile):
    prefix = "" if section in profile["base_sections"] else f"{slugify(section)}-"
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", f"{prefix}{number}").strip("-")
    return f"{safe}.yaml"


def display_section(section, profile):
    return "Base" if section in profile["base_sections"] else section


def card_title(card, profile):
    return " / ".join(subject["name"] for subject in card["subjects"])


def card_name(card, profile):
    title = card_title(card, profile)
    section = display_section(card["section"], profile)
    return f"{title} - {section}"


def card_description(card, profile):
    title = card_title(card, profile)
    section = display_section(card["section"], profile)
    if is_autograph(card, profile):
        if section.lower().startswith("autographed"):
            return f"{section} card featuring {title}."
        return f"Autographed {section} card featuring {title}."
    if is_relic(card, profile):
        return f"{section} memorabilia card featuring {title}."
    return f"{section} card featuring {title}."


def is_relic(card, profile):
    if card["section"] in profile.get("section_relics", set()):
        return True
    section = card["section"].lower()
    return any(term in section for term in ["relic", "material", "swatch", "patch", "jersey", "memorabilia"])


def is_autograph(card, profile):
    section = card["section"].lower()
    return any(term in section for term in ["autograph", "signature", "signatures", "aurograph"])


def variation_value(section):
    normalized = section.lower()
    if "variation" not in normalized:
        return None
    if "red rc" in normalized:
        return "Red RC Variation"
    if "etched in glass" in normalized:
        return "Etched In Glass Variation"
    if "city" in normalized:
        return "City Variation"
    if "image" in normalized:
        return "Image Variation"
    if "packfractor" in normalized:
        return "Packfractor Variation"
    if "kanji" in normalized:
        return "Kanji Variation"
    return section


def write_set(set_dir, set_id, source_url, source_file, description, cards, profile):
    set_path = set_dir / "set.yaml"
    counts = {}
    for card in cards:
        counts[card["section"]] = counts.get(card["section"], 0) + 1

    lines = [
        f"uuid: {get_uuid(set_path)}",
        f"set_id: {yaml_quote(set_id)}",
        f"name: {yaml_quote(profile['set_name'])}",
        'genre: "Sports"',
        f"category: [{', '.join(yaml_quote(category) for category in profile.get('category', ['MLB']))}]",
        'sports: ["Baseball"]',
        f"season: {yaml_quote(profile['season'])}",
        f"years: [{', '.join(str(year) for year in profile['years'])}]",
        f"series: {yaml_quote(profile['series'])}",
    ]
    if "series_number" in profile:
        lines.append(f"series_number: {profile['series_number']}")
    lines.extend(
        [
            f"release_date: {yaml_quote(profile['release_date'])}",
            f"manufacturer: {yaml_quote(profile['manufacturer'])}",
            f"card_count: {len(cards)}",
            f"description: {yaml_quote(description)}",
            f"image_url: {yaml_quote(profile['set_image_url'])}",
            "metadata:",
            '  language: "en"',
            f"  year: {profile['years'][-1]}",
            f"  source_name: {yaml_quote('BaseballCardpedia')}",
            f"  source_url: {yaml_quote(source_url)}",
            f"  source_file: {yaml_quote(source_file)}",
            "  subsets:",
        ]
    )
    for section, count in counts.items():
        lines.append(f"    - name: {yaml_quote(section)}")
        lines.append(f"      card_count: {count}")

    set_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_cards(set_dir, set_id, cards, profile):
    cards_dir = set_dir / "cards"
    cards_dir.mkdir(parents=True, exist_ok=True)
    section_print_runs = profile.get("section_print_runs", {})
    section_estimated_print_runs = profile.get("section_estimated_print_runs", {})

    for card in cards:
        path = cards_dir / card_filename(card["number"], card["section"], profile)
        section = display_section(card["section"], profile)
        autograph = is_autograph(card, profile)
        relic = is_relic(card, profile)
        role = "Player"
        if "Team Card" in card.get("note", ""):
            role = "Team"
        image_slug = path.stem
        print_run = card.get("print_run") or section_print_runs.get(card["section"])
        if print_run is None and profile.get("all_cards_serial_numbered"):
            print_run = profile.get("default_print_run")
        estimated_print_run = section_estimated_print_runs.get(card["section"])
        serial_numbered = bool(card.get("serial_numbered")) or print_run is not None

        lines = [
            f"number: {yaml_quote(card['number'])}",
            f"uuid: {yaml_quote(get_uuid(path))}",
            'genre: "Sports"',
            'sport: "Baseball"',
            f"set_id: {yaml_quote(set_id)}",
            "subjects:",
        ]
        for subject in card["subjects"]:
            lines.append(f"  - name: {yaml_quote(subject['name'])}")
            lines.append(f"    role: {yaml_quote(role)}")
            lines.append(f"    team: {yaml_quote(subject['team'])}")

        if card["section"] not in profile["base_sections"]:
            lines.append(f"subset: {yaml_quote(card['section'])}")
        lines.extend(
            [
                f"card_name: {yaml_quote(card_name(card, profile))}",
                f"description: {yaml_quote(card_description(card, profile))}",
                f"series: {yaml_quote(profile['series'])}",
            ]
        )
        variation = variation_value(card["section"])
        if variation:
            lines.append(f"variation: {yaml_quote(variation)}")
        if print_run is not None:
            lines.append(f"print_run: {print_run}")
        elif estimated_print_run is not None:
            lines.append(f"print_run: {estimated_print_run}")
        lines.extend(
            [
                f"rookie_card: {str(card['rookie_card']).lower()}",
                f"autograph: {str(autograph).lower()}",
                f"relic: {str(relic).lower()}",
                f"serial_numbered: {str(serial_numbered).lower()}",
                f"release_date: {yaml_quote(profile['release_date'])}",
                f"image_url: {yaml_quote(f'{profile['image_root']}/{image_slug}.jpg')}",
            ]
        )

        path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--set-id", choices=sorted(PROFILES), default="2025-26-topps-chrome-platinum")
    parser.add_argument("--url")
    parser.add_argument("--xlsx")
    parser.add_argument("--category", default="baseball")
    return parser.parse_args()


def main():
    args = parse_args()
    profile = PROFILES[args.set_id]
    source_url = args.url or profile["url"]
    source_file = args.xlsx or profile["xlsx"]
    cards = read_cards_from_workbook(Path(source_file), profile)
    set_dir = DATA_DIR / args.category / args.set_id
    set_dir.mkdir(parents=True, exist_ok=True)
    description = fetch_description(source_url, f"{profile['set_name']} checklist.")
    write_set(set_dir, args.set_id, source_url, source_file, description, cards, profile)
    write_cards(set_dir, args.set_id, cards, profile)
    print(f"Wrote {len(cards)} cards to {set_dir}")


if __name__ == "__main__":
    main()
